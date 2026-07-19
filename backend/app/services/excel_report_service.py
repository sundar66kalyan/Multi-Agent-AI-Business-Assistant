from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelReportService:

    @staticmethod
    def generate(report_data: dict):

        finance = report_data["finance"]
        analytics = report_data["analytics"]

        reports = Path("reports")
        reports.mkdir(exist_ok=True)

        filename = (
            f"Business_Report_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        output = reports / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "Business Report"

        # Title
        ws["A1"] = "MULTI-AGENT AI BUSINESS ASSISTANT"
        ws["A1"].font = Font(size=16, bold=True)

        ws["A3"] = "Finance Summary"
        ws["A3"].font = Font(bold=True)

        ws.append(["Metric", "Value"])
        ws.append(["Revenue", finance["revenue"]])
        ws.append(["Expenses", finance["expenses"]])
        ws.append(["Profit", finance["profit"]])
        ws.append(["Month", finance["month"]])

        ws.append([])
        ws.append(["System Analytics"])
        ws.append(["Metric", "Value"])
        ws.append(["Indexed Documents", analytics["documents"]])
        ws.append(["Chunks", analytics["chunks"]])
        ws.append(["Finance Records", analytics["finance_records"]])

        wb.save(output)

        return str(output)